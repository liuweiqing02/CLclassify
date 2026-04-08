import json
import os
import random
import re
from dataclasses import dataclass
from multiprocessing import Manager
from typing import Dict, List, Tuple

import nibabel as nib
import numpy as np
import torch
from scipy.ndimage import zoom
from torch.utils.data import Dataset


@dataclass
class DongmaiRecord:
    patient_id: str
    ct_image: str
    ct_mask: str
    mri_image: str
    mri_mask: str
    label: int


def _strip_nii_ext(filename: str) -> str:
    if filename.endswith(".nii.gz"):
        return filename[:-7]
    if filename.endswith(".nii"):
        return filename[:-4]
    return os.path.splitext(filename)[0]


def canonical_patient_id(name: str) -> str:
    base = _strip_nii_ext(os.path.basename(str(name))).lower()

    suffixes = [
        "_0000", "_0001", "_0002", "_0003",
        "_ct", "_mr", "_mri", "-ct", "-mr", "-mri",
        "_image", "_img", "_label", "_labels", "_seg", "_mask",
    ]
    changed = True
    while changed:
        changed = False
        for s in suffixes:
            if base.endswith(s):
                base = base[: -len(s)]
                changed = True

    digits = re.findall(r"\d+", base)
    if digits:
        return str(int(digits[-1]))
    return base


def _read_excel_labels(excel_path: str) -> Dict[str, int]:
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"Class label file not found: {excel_path}")

    try:
        from openpyxl import load_workbook  # type: ignore

        wb = load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise ValueError(f"Excel file is empty: {excel_path}")

        header = [str(x).strip().lower() if x is not None else "" for x in rows[0]]

        def _find_col(candidates: Tuple[str, ...], default_idx: int) -> int:
            for i, h in enumerate(header):
                if any(k in h for k in candidates):
                    return i
            return default_idx

        id_col = _find_col(("id", "patient"), 0)
        cls_col = _find_col(("class", "label", "类别", "分型"), 1 if len(header) > 1 else 0)

        raw = {}
        for row in rows[1:]:
            if row is None:
                continue
            rid = row[id_col] if id_col < len(row) else None
            rcls = row[cls_col] if cls_col < len(row) else None
            if rid is None or rcls is None:
                continue
            pid = canonical_patient_id(str(rid))
            try:
                cls_val = int(rcls)
            except (ValueError, TypeError):
                continue
            raw[pid] = cls_val
        return raw
    except Exception:
        pass

    try:
        import pandas as pd  # type: ignore

        df = pd.read_excel(excel_path)
        cols = [str(c).strip().lower() for c in df.columns]

        id_col_name = df.columns[0]
        cls_col_name = df.columns[1] if len(df.columns) > 1 else df.columns[0]
        for i, c in enumerate(cols):
            if "id" in c or "patient" in c:
                id_col_name = df.columns[i]
            if "class" in c or "label" in c or "类别" in c or "分型" in c:
                cls_col_name = df.columns[i]

        raw = {}
        for _, row in df.iterrows():
            if row[id_col_name] is None or row[cls_col_name] is None:
                continue
            pid = canonical_patient_id(str(row[id_col_name]))
            try:
                cls_val = int(row[cls_col_name])
            except (ValueError, TypeError):
                continue
            raw[pid] = cls_val
        return raw
    except Exception as e:
        raise RuntimeError(
            f"Failed to read class labels from {excel_path}. "
            f"Please install openpyxl or pandas. Details: {e}"
        )


def _build_file_map(folder: str) -> Dict[str, str]:
    files = [f for f in os.listdir(folder) if f.endswith(".nii") or f.endswith(".nii.gz")]
    file_map = {}
    for f in sorted(files):
        pid = canonical_patient_id(f)
        file_map[pid] = os.path.join(folder, f)
    return file_map


def build_dongmai_records(config) -> Tuple[List[DongmaiRecord], Dict[int, int]]:
    ct_img_map = _build_file_map(config.image_dir_CT_tr)
    ct_mask_map = _build_file_map(config.label_dir_CT_tr)
    mri_img_map = _build_file_map(config.image_dir_MRI_tr)
    mri_mask_map = _build_file_map(config.label_dir_MRI_tr)

    label_raw = _read_excel_labels(config.class_label_path)

    valid_ids = set(ct_img_map) & set(ct_mask_map) & set(mri_img_map) & set(mri_mask_map) & set(label_raw)
    if not valid_ids:
        raise RuntimeError(
            "No matched patient ids found across CT/MRI images, masks and class labels. "
            "Please check ID naming consistency."
        )

    unique_src_labels = sorted({label_raw[pid] for pid in valid_ids})
    label_encoder = {src: i for i, src in enumerate(unique_src_labels)}

    records = []
    for pid in sorted(valid_ids, key=lambda x: (len(x), x)):
        records.append(
            DongmaiRecord(
                patient_id=pid,
                ct_image=ct_img_map[pid],
                ct_mask=ct_mask_map[pid],
                mri_image=mri_img_map[pid],
                mri_mask=mri_mask_map[pid],
                label=label_encoder[label_raw[pid]],
            )
        )

    label_decode = {v: k for k, v in label_encoder.items()}
    return records, label_decode


def split_records_patient_level(
    records: List[DongmaiRecord],
    val_ratio: float = 0.2,
    seed: int = 42,
) -> Tuple[List[DongmaiRecord], List[DongmaiRecord]]:
    if not 0.0 <= val_ratio < 1.0:
        raise ValueError("val_ratio must be in [0, 1)")
    if val_ratio >= 1.0:
        raise ValueError("val_ratio must be < 1")

    rng = random.Random(seed)
    by_label: Dict[int, List[DongmaiRecord]] = {}
    for r in records:
        by_label.setdefault(r.label, []).append(r)

    train, val = [], []
    for _, group in by_label.items():
        group = group[:]
        rng.shuffle(group)

        n = len(group)
        n_val = int(round(n * val_ratio))

        if n >= 2 and n_val >= n:
            n_val = n - 1

        val.extend(group[:n_val])
        train.extend(group[n_val:])

    rng.shuffle(train)
    rng.shuffle(val)
    return train, val


def save_split_file(
    path: str,
    train_records: List[DongmaiRecord],
    val_records: List[DongmaiRecord],
    test_records: List[DongmaiRecord] = None,
):
    if test_records is None:
        test_records = []
    payload = {
        "train_ids": [r.patient_id for r in train_records],
        "val_ids": [r.patient_id for r in val_records],
        "test_ids": [r.patient_id for r in test_records],
    }
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, ensure_ascii=False)


def load_split_file(path: str) -> Dict[str, List[str]]:
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    for k in ("train_ids", "val_ids"):
        if k not in data:
            raise ValueError(f"Split file missing key: {k}")
    if "test_ids" not in data:
        data["test_ids"] = []
    return data


def split_records_by_ids(
    records: List[DongmaiRecord],
    split_data: Dict[str, List[str]],
) -> Tuple[List[DongmaiRecord], List[DongmaiRecord]]:
    rec_map = {r.patient_id: r for r in records}

    # Backward compatibility: if an old split contains test_ids, merge them into training.
    train_ids = list(split_data["train_ids"]) + list(split_data.get("test_ids", []))
    train = [rec_map[pid] for pid in train_ids if pid in rec_map]
    val = [rec_map[pid] for pid in split_data["val_ids"] if pid in rec_map]

    if not train or not val:
        raise RuntimeError("Loaded split produced empty train/val. Please regenerate split file.")
    return train, val


def _load_nifti(path: str) -> Tuple[np.ndarray, Tuple[float, float, float]]:
    nii = nib.load(path)
    data = nii.get_fdata().astype(np.float32)
    spacing = tuple(float(s) for s in nii.header.get_zooms()[:3])
    return data, spacing


def _resample(data: np.ndarray, original_spacing, desired_spacing, is_label: bool) -> np.ndarray:
    resize_factor = np.array(original_spacing, dtype=np.float32) / np.array(desired_spacing, dtype=np.float32)
    new_shape = np.maximum(1, np.round(np.array(data.shape) * resize_factor)).astype(np.int32)
    real_factor = new_shape / np.array(data.shape)
    order = 0 if is_label else 1
    return zoom(data, real_factor, order=order, mode="nearest")


def _pad_or_crop_center(data: np.ndarray, target_shape: Tuple[int, int, int]) -> np.ndarray:
    out = data
    for axis, target in enumerate(target_shape):
        cur = out.shape[axis]
        if cur < target:
            pad_before = (target - cur) // 2
            pad_after = target - cur - pad_before
            pad_width = [(0, 0), (0, 0), (0, 0)]
            pad_width[axis] = (pad_before, pad_after)
            out = np.pad(out, pad_width, mode="constant", constant_values=0)
        elif cur > target:
            start = (cur - target) // 2
            end = start + target
            slicer = [slice(None), slice(None), slice(None)]
            slicer[axis] = slice(start, end)
            out = out[tuple(slicer)]
    return out


def _zscore_norm(img: np.ndarray) -> np.ndarray:
    m = float(img.mean())
    s = float(img.std())
    if s < 1e-6:
        return img - m
    return (img - m) / s


class DongmaiPairDataset(Dataset):
    _manager = None
    _shared_cache = None
    _lock = None

    def __init__(self, records: List[DongmaiRecord], config, training: bool = False):
        self.records = records
        self.config = config
        self.training = training
        self.desired_spacing = config.desired_spacing
        self.target_size = config.target_size
        self.target_size_crop = config.target_size_crop
        self.use_cache = True

        if DongmaiPairDataset._manager is None:
            DongmaiPairDataset._manager = Manager()
            DongmaiPairDataset._shared_cache = DongmaiPairDataset._manager.dict()
            DongmaiPairDataset._lock = DongmaiPairDataset._manager.Lock()

        self.cache = DongmaiPairDataset._shared_cache
        self.cache_lock = DongmaiPairDataset._lock

    def __len__(self):
        return len(self.records)

    def _preprocess_volume(self, path: str, is_label: bool) -> np.ndarray:
        data, spacing = _load_nifti(path)
        if is_label:
            data = (data > 0).astype(np.float32)

        data = _resample(data, spacing, self.desired_spacing, is_label=is_label)
        data = _pad_or_crop_center(data, self.target_size)
        data = _pad_or_crop_center(data, self.target_size_crop)

        if not is_label:
            data = _zscore_norm(data)
        return data.astype(np.float32)

    def _augment_pair(self, ct: np.ndarray, mri: np.ndarray) -> Tuple[np.ndarray, np.ndarray]:
        if not self.training:
            return ct, mri

        flip_flags = [random.random() < 0.5 for _ in range(3)]
        do_rot = random.random() < 0.5
        rot_axes_choices = [(1, 2), (1, 3), (2, 3)]  # [C, D, H, W]
        rot_axes = random.choice(rot_axes_choices)
        rot_k = random.randint(0, 3)

        do_crop = random.random() < 0.5
        crop_ratio = random.uniform(0.8, 1.0)
        crop_window = None
        if do_crop:
            _, d, h, w = ct.shape
            cd = max(2, int(d * crop_ratio))
            ch = max(2, int(h * crop_ratio))
            cw = max(2, int(w * crop_ratio))
            sd = random.randint(0, max(0, d - cd))
            sh = random.randint(0, max(0, h - ch))
            sw = random.randint(0, max(0, w - cw))
            crop_window = (sd, sh, sw, cd, ch, cw)

        def _resize_3d(vol: np.ndarray, target_shape: Tuple[int, int, int], is_label: bool) -> np.ndarray:
            factor = np.array(target_shape, dtype=np.float32) / np.array(vol.shape, dtype=np.float32)
            return zoom(vol, factor, order=0 if is_label else 1, mode="nearest").astype(np.float32)

        def _apply_geom(branch: np.ndarray) -> np.ndarray:
            x = branch

            for i, f in enumerate(flip_flags):
                if f:
                    x = np.flip(x, axis=i + 1).copy()

            if do_rot and rot_k > 0:
                x = np.rot90(x, k=rot_k, axes=rot_axes).copy()

            if do_crop and crop_window is not None:
                _, d, h, w = x.shape
                sd, sh, sw, cd, ch, cw = crop_window
                cropped = x[:, sd:sd + cd, sh:sh + ch, sw:sw + cw]
                out = np.empty_like(x, dtype=np.float32)
                out[0] = _resize_3d(cropped[0], (d, h, w), is_label=False)
                out[1] = _resize_3d(cropped[1], (d, h, w), is_label=True)
                x = out

            return x

        ct = _apply_geom(ct)
        mri = _apply_geom(mri)

        if random.random() < 0.3:
            scale = random.uniform(0.9, 1.1)
            bias = random.uniform(-0.1, 0.1)
            ct[0] = ct[0] * scale + bias
            mri[0] = mri[0] * scale + bias

        return ct, mri

    def __getitem__(self, index: int):
        rec = self.records[index]
        cache_key = rec.patient_id

        if self.use_cache and cache_key in self.cache:
            ct_branch_cached, mri_branch_cached = self.cache[cache_key]
            ct_branch = ct_branch_cached.copy()
            mri_branch = mri_branch_cached.copy()
        else:
            ct_img = self._preprocess_volume(rec.ct_image, is_label=False)
            ct_mask = self._preprocess_volume(rec.ct_mask, is_label=True)
            mri_img = self._preprocess_volume(rec.mri_image, is_label=False)
            mri_mask = self._preprocess_volume(rec.mri_mask, is_label=True)

            ct_branch = np.stack([ct_img, ct_mask], axis=0).astype(np.float32)
            mri_branch = np.stack([mri_img, mri_mask], axis=0).astype(np.float32)

            if self.use_cache:
                with self.cache_lock:
                    if cache_key not in self.cache:
                        self.cache[cache_key] = (ct_branch.copy(), mri_branch.copy())

        ct_branch, mri_branch = self._augment_pair(ct_branch, mri_branch)

        return (
            torch.from_numpy(ct_branch),
            torch.from_numpy(mri_branch),
            torch.tensor(rec.label, dtype=torch.long),
            rec.patient_id,
        )

    @staticmethod
    def collate_fn(batch):
        ct = torch.stack([item[0] for item in batch], dim=0)
        mri = torch.stack([item[1] for item in batch], dim=0)
        y = torch.stack([item[2] for item in batch], dim=0)
        ids = [item[3] for item in batch]
        return ct, mri, y, ids
